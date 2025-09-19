import streamlit as st
import pandas as pd
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.chart import BarChart, Reference

# File penyimpanan data
DATA_FILE = "data_progress.csv"
LAPORAN_FILE = "data_laporan.csv"

# Inisialisasi data jika belum ada
if not os.path.exists(DATA_FILE):
    pd.DataFrame(columns=["Blok", "Tanggal", "Item", "Prosentase", "Nilai SPK"]).to_csv(DATA_FILE, index=False)
if not os.path.exists(LAPORAN_FILE):
    pd.DataFrame(columns=["Blok", "Tanggal", "Item", "Prosentase", "Nilai SPK", "Tanggal_Laporan"]).to_csv(LAPORAN_FILE, index=False)

# Fungsi untuk load data
def load_data(file):
    return pd.read_csv(file)

# Fungsi untuk simpan data
def save_data(df, file):
    df.to_csv(file, index=False)

st.title("📊 Aplikasi Monitoring Pembangunan Perumahan")
st.write("Input progress tiap blok rumah, tandai sudah dilaporkan, dan buat laporan rekap.")

# --- Form Input ---
st.header("➕ Input Progress Baru")
with st.form("input_form"):
    blok = st.text_input("Nama Blok")
    tanggal = st.date_input("Tanggal", datetime.today())
    item = st.text_input("Item Pekerjaan")
    prosentase = st.number_input("Prosentase (%)", 0, 100, 0)
    nilai_spk = st.number_input("Nilai SPK (Rp)", 0, step=1000000)
    submit = st.form_submit_button("Simpan")

if submit:
    if blok and item:
        df = load_data(DATA_FILE)
        new_data = pd.DataFrame([[blok, tanggal, item, prosentase, nilai_spk]],
                                columns=["Blok", "Tanggal", "Item", "Prosentase", "Nilai SPK"])
        df = pd.concat([df, new_data], ignore_index=True)
        save_data(df, DATA_FILE)
        st.success("✅ Data berhasil disimpan!")
    else:
        st.error("❌ Harap isi Nama Blok dan Item Pekerjaan")

# --- Tabel Data Belum Dilaporkan ---
st.header("📋 Data Belum Dilaporkan")
data = load_data(DATA_FILE)
if not data.empty:
    for i, row in data.iterrows():
        col1, col2 = st.columns([4,1])
        with col1:
            st.write(f"**{row['Blok']}** | {row['Tanggal']} | {row['Item']} | {row['Prosentase']}% | Rp {row['Nilai SPK']:,}")
        with col2:
            if st.button("Sudah Dilaporkan", key=f"lapor_{i}"):
                laporan = load_data(LAPORAN_FILE)
                row_dict = row.to_dict()
                row_dict["Tanggal_Laporan"] = datetime.today().strftime("%Y-%m-%d")
                laporan = pd.concat([laporan, pd.DataFrame([row_dict])], ignore_index=True)
                save_data(laporan, LAPORAN_FILE)

                # Hapus dari data utama
                data = data.drop(i)
                save_data(data, DATA_FILE)
                st.experimental_rerun()
else:
    st.info("Belum ada data progress baru.")

# --- Laporan ---
st.header("📑 Laporan Progress")
laporan = load_data(LAPORAN_FILE)
if not laporan.empty:
    st.dataframe(laporan)
    
    # Rekap per blok
    rekap = laporan.groupby("Blok")["Prosentase"].mean().reset_index()
    st.subheader("📈 Rekap Rata-rata Progress per Blok")
    st.bar_chart(rekap.set_index("Blok"))

    # Download laporan Excel dengan format blok terpisah + styling + grafik
    def export_excel_by_block(df):
        wb = Workbook()
        ws = wb.active
        ws.title = "Laporan"

        # Style dasar
        bold_font = Font(bold=True)
        header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        border_style = Border(left=Side(style="thin"),
                              right=Side(style="thin"),
                              top=Side(style="thin"),
                              bottom=Side(style="thin"))

        # Judul
        ws["A1"] = "LAPORAN PROGRESS PEMBANGUNAN"
        ws["A1"].font = Font(size=14, bold=True)
        ws["A2"] = "Per Tanggal: " + datetime.today().strftime("%Y-%m-%d")

        row_start = 4
        total_list = []
        rata_per_blok = []

        for blok, group in df.groupby("Blok"):
            ws[f"A{row_start}"] = f"BLOK {blok}"
            ws[f"A{row_start}"].font = Font(bold=True, size=12)
            row_start += 1

            # Header tabel
            headers = ["NO", "TANGGAL", "ITEM PEKERJAAN", "PROSENTASE", "NILAI SPK"]
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=row_start, column=col, value=header)
                cell.font = bold_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border_style
            row_start += 1

            # Data per item pekerjaan
            for i, r in enumerate(group.itertuples(index=False), start=1):
                ws.cell(row=row_start, column=1, value=i).border = border_style
                ws.cell(row=row_start, column=2, value=str(r.Tanggal)).border = border_style
                ws.cell(row=row_start, column=3, value=r.Item).border = border_style
                ws.cell(row=row_start, column=4, value=f"{r.Prosentase}%").border = border_style
                ws.cell(row=row_start, column=5, value=f"Rp {int(r._4):,}").border = border_style if len(r) > 4 else ""
                row_start += 1

            # Subtotal rata-rata per blok
            rata2 = group["Prosentase"].mean()
            ws[f"A{row_start}"] = f"Subtotal Blok {blok} (Rata-rata Progres):"
            ws[f"D{row_start}"] = f"{rata2:.2f}%"
            ws[f"A{row_start}"].font = Font(italic=True)
            ws[f"D{row_start}"].font = Font(bold=True)
            total_list.append(rata2)
            rata_per_blok.append((blok, rata2))
            row_start += 2

        # Total rata-rata semua blok
        ws[f"A{row_start}"] = "TOTAL Rata-rata Semua Blok:"
        ws[f"D{row_start}"] = f"{pd.Series(total_list).mean():.2f}%"
        ws[f"A{row_start}"].font = Font(bold=True)
        ws[f"D{row_start}"].font = Font(bold=True)

        # Tambahkan grafik progres per blok di halaman paling akhir
        if rata_per_blok:
            chart_row = row_start + 4
            ws[f"A{chart_row}"] = "Grafik Rata-rata Progress per Blok"
            ws[f"A{chart_row}"].font = Font(bold=True)
            chart_row += 1

            # Tulis data grafik
            ws.cell(row=chart_row, column=1, value="Blok").font = bold_font
            ws.cell(row=chart_row, column=2, value="Progres").font = bold_font
            chart_row += 1

            for i, (blok, val) in enumerate(rata_per_blok, start=0):
                ws.cell(row=chart_row+i, column=1, value=blok)
                ws.cell(row=chart_row+i, column=2, value=val)

            chart = BarChart()
            data_ref = Reference(ws, min_col=2, min_row=chart_row, max_row=chart_row+len(rata_per_blok)-1)
            cats_ref = Reference(ws, min_col=1, min_row=chart_row, max_row=chart_row+len(rata_per_blok)-1)
            chart.add_data(data_ref, titles_from_data=False)
            chart.set_categories(cats_ref)
            chart.title = "Rata-rata Progress per Blok"
            chart.y_axis.title = "% Progres"
            chart.x_axis.title = "Blok"
            ws.add_chart(chart, f"E{chart_row}")

        return wb

    excel_file = "laporan_progress.xlsx"
    wb = export_excel_by_block(laporan)
    wb.save(excel_file)

    with open(excel_file, "rb") as f:
        st.download_button("📥 Download Laporan (Excel)", f, excel_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

else:
    st.info("Belum ada laporan.")
